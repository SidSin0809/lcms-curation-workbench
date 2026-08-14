from __future__ import annotations

import os
import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from lcms_curation.engine import FILE_ROLES, analyze_files, apply_filters, evaluate_filters
from lcms_curation.exports import export_workbook
from lcms_curation.provenance import score_provenance
from lcms_curation.qc import default_qc_settings
from lcms_curation.reporting import journal_report_markdown


class FullDatasetValidation(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        source = os.environ.get("LCMS_TEST_DATA")
        if not source:
            raise unittest.SkipTest("Set LCMS_TEST_DATA to a folder containing CM.csv, AM.csv, CI.csv, II.xml, FD.msp, and ACP.csv")
        cls.files = {role: Path(source) / f"{role}.{'xml' if role == 'II' else 'msp' if role == 'FD' else 'csv'}" for role in FILE_ROLES}
        cls.bundle = analyze_files(cls.files, {"ion_mode": "negative", "mass_tolerance": 10.0, "qc_pattern": "QC"})

    def test_known_dataset_signature(self) -> None:
        self.assertEqual(len(self.bundle.accepted_reads), 1375)
        self.assertEqual(len(self.bundle.ci), 38850)
        self.assertEqual(len(self.bundle.selected_reads), 1091)
        self.assertEqual(self.bundle.duplicate_groups, 201)
        self.assertEqual(len(self.bundle.duplicate_reads), 485)
        self.assertEqual(self.bundle.extra_reads_collapsed, 284)

    def test_balanced_thresholds(self) -> None:
        balanced = next(item for item in self.bundle.thresholds["presets"] if item["name"] == "Balanced")
        self.assertEqual(balanced["score"], 35.8)
        self.assertEqual(balanced["fragmentation"], 2.9)
        self.assertEqual(balanced["abs_mass_error"], 4.8)
        self.assertEqual(balanced["isotope"], 77.9)
        shortlisted = apply_filters(self.bundle.selected_reads, balanced)
        self.assertEqual(len(shortlisted), 315)

    def test_sample_ms2_chemistry_and_qc_contract(self) -> None:
        roles = self.bundle.sample_metadata["Sample role"].value_counts().to_dict()
        self.assertEqual(roles.get("Biological"), 32)
        self.assertEqual(roles.get("Pooled QC"), 5)
        self.assertEqual(int(self.bundle.fd["FD observed peaks"].sum()), len(self.bundle.fd_peaks))
        self.assertIn("Relative intensity (%)", self.bundle.fd_peaks.columns)
        self.assertIn("Neutral loss from precursor (Da)", self.bundle.fd_peaks.columns)
        self.assertIn("FD normalized spectral entropy", self.bundle.fd.columns)
        self.assertIn("Formula DBE", self.bundle.selected_reads.columns)
        self.assertIn("Formula polarity proxy (0-100)", self.bundle.selected_reads.columns)
        self.assertEqual(int(self.bundle.selected_reads["Formula parse status"].eq("Parsed").sum()), 1091)
        self.assertIn("QC robust CV% (MAD)", self.bundle.analytical_qc.columns)
        self.assertIn("QC drift across run (%)", self.bundle.analytical_qc.columns)
        self.assertIn("D-ratio (%)", self.bundle.analytical_qc.columns)
        self.assertEqual(len(self.bundle.analytical_qc), 1091)

    def test_combined_filter_decisions_are_auditable(self) -> None:
        balanced = next(item for item in self.bundle.thresholds["presets"] if item["name"] == "Balanced")
        decisions = evaluate_filters(self.bundle.selected_reads, {**balanced, **default_qc_settings()})
        self.assertEqual(len(decisions), 1091)
        self.assertEqual(int(decisions["Identification evidence pass"].sum()), 315)
        self.assertIn("Filter fail reasons", decisions.columns)
        self.assertIn("Analytical QC evidence available", decisions.columns)

    def test_context_and_workbook_export(self) -> None:
        balanced = next(item for item in self.bundle.thresholds["presets"] if item["name"] == "Balanced")
        filters = {
            **balanced,
            **default_qc_settings(),
            "apply_drift_filter": True,
            "apply_d_ratio_filter": True,
        }
        shortlisted = apply_filters(self.bundle.selected_reads, filters)
        self.assertEqual(len(shortlisted), 214)
        context = {
            "study_type": "lipidomics",
            "assay_type": "lipidomics",
            "biological_system": "human",
            "sample_matrix": "Tissue",
            "exposure_context": "none",
            "extraction_method": "mtbe",
            "analyzed_phase": "upper-organic",
        }
        provenance = score_provenance(shortlisted, context)
        self.assertEqual(len(provenance), 214)
        self.assertIn("Primary source class", provenance.columns)
        report = journal_report_markdown(self.bundle, shortlisted, provenance, filters, context)
        self.assertIn("1,091 accepted entities", report)
        self.assertIn("214 shortlisted compounds", report)
        self.assertIn("not a target–decoy-controlled FDR", report)
        self.assertIn("Information not recoverable from the six exports", report)
        tasks = pd.DataFrame(
            [{
                "Stage": "Step 07 · Context model",
                "Task": "Score contextual evidence",
                "Status": "Completed",
                "Started": "2026-08-13T21:00:00+05:30",
                "Finished": "2026-08-13T21:00:01+05:30",
                "Duration (s)": 1.0,
                "Outcome": "214 compounds scored",
            }]
        )
        with tempfile.TemporaryDirectory() as directory:
            target = Path(directory) / "validation.xlsx"
            export_workbook(
                self.bundle,
                shortlisted,
                provenance,
                filters,
                context,
                target,
                include_all_candidates=False,
                task_history=tasks,
            )
            self.assertTrue(target.is_file())
            self.assertGreater(target.stat().st_size, 100_000)
            workbook = load_workbook(target, read_only=True)
            for sheet in (
                "00 Read Me",
                "02 Sample Metadata",
                "06 MS2 Spectra",
                "07 MS2 Peaks",
                "09 Analytical QC",
                "11 Filter Decisions",
                "14 Journal Methods",
                "15 Task History",
            ):
                self.assertIn(sheet, workbook.sheetnames)
            workbook.close()


if __name__ == "__main__":
    unittest.main()
