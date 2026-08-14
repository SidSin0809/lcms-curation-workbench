from __future__ import annotations

import json
import math
from collections.abc import Callable
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .engine import AnalysisBundle, evaluate_filters
from .metadata import metadata_requirement_table, project_metadata_frame
from .reporting import journal_report_markdown, journal_report_table
from . import __version__

NAVY = "13284A"
BLUE = "1F5AE0"
PALE_BLUE = "DDE9FF"
PALE_AQUA = "DDF5F2"
INK = "182B47"
MUTED = "657188"


def _emit(progress: Callable[[int, str], None] | None, percent: int, message: str) -> None:
    if progress is not None:
        progress(max(0, min(100, int(percent))), message)


def _safe_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (float, np.floating)):
        return None if not np.isfinite(value) else float(value)
    if isinstance(value, (int, np.integer)):
        return int(value)
    if isinstance(value, (bool, np.bool_)):
        return bool(value)
    text = str(value)
    return f"'{text}" if text.startswith(("=", "+", "-", "@")) else text


def safe_frame(frame: pd.DataFrame) -> pd.DataFrame:
    output = frame.copy()
    for column in output.columns:
        if output[column].dtype == object or pd.api.types.is_string_dtype(output[column].dtype):
            output[column] = output[column].map(_safe_value)
    return output.replace({np.nan: None, np.inf: None, -np.inf: None})


def export_csv(frame: pd.DataFrame, path: str | Path) -> Path:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    safe_frame(frame).to_csv(target, index=False, encoding="utf-8-sig", lineterminator="\r\n")
    return target


def threshold_table(bundle: AnalysisBundle) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for key, label, direction in (
        ("score", "Score", ">="),
        ("fragmentation", "Fragmentation Score", ">="),
        ("abs_mass_error", "Absolute Mass Error (ppm)", "<="),
        ("isotope", "Isotope Similarity", ">="),
    ):
        field = bundle.thresholds[key]
        rows.append(
            {
                "Metric": label,
                "Recommended cutoff": field["value"],
                "Direction": direction,
                "Bootstrap CI low": field["ci_low"],
                "Bootstrap CI high": field["ci_high"],
                "Method": field["method"],
                "Rationale": field["rationale"],
            }
        )
    return pd.DataFrame(rows)


def method_notes(bundle: AnalysisBundle) -> pd.DataFrame:
    notes = [
        ("Input reconciliation", "All six sources are joined by the Progenesis feature identifier. Source-row counts, duplicate keys, CM coverage, SHA-256 hashes, and structural findings are retained."),
        ("Sample roles", "Sample roles are mapped explicitly from the CM header and remain user-editable. Pooled QC, biological samples, blanks, reference material, calibration, system suitability, and exclusions are not interchangeable."),
        ("Pooled QC", "QC detection, conventional CV, robust MAD-based CV, D-ratio, technical-replicate CV, and injection-order drift are reported when the required mapped samples exist. Missing QC or blank evidence is marked not evaluable."),
        ("Entity key", "Accepted Compound ID; fallback to Formula + canonical name; final fallback to feature ID."),
        ("Representative read", "Lexicographic: Score descending, Fragmentation Score descending, absolute mass error ascending, Isotope Similarity descending; then evidence completeness, spectrum richness, QC detection/CV, biological detection/abundance, and stable Feature ID."),
        ("Repeated identities", "Repeated Accepted Compound IDs may reflect isomers, in-source products, adduct behavior, or false matches. The separate repeated-read file must be retained."),
        ("Thresholds", "Entity-level robust quantiles with 500 deterministic bootstrap resamples. A stable one-dimensional two-cluster mass-error boundary is used only when Gaussian validation has delta-BIC > 10 and Ashman D >= 2, capped by search tolerance."),
        ("Fragmentation zero", "A zero is treated as an explicit unavailable/disabled/unmatched evidence state in Inclusive and Balanced presets; Stringent requires positive fragmentation support."),
        ("MS2 extraction", "Every MSP spectrum and peak is retained. Spectral entropy, normalized entropy, effective peak count, base peak, low-intensity share, and peak-count integrity describe complexity; without a reference/decoy library they do not establish identity or FDR."),
        ("Chemical properties", "Element counts, exact formula mass, DBE, H/C, O/C, N/C, formula-polarity proxy, lipid-like flag, and Kendrick mass defect are formula-derived. They do not encode isomerism, pKa, measured logP, or measured recovery."),
        ("Threshold limitation", "Dataset-adaptive triage, not a target-decoy FDR or confirmed-identification boundary."),
        ("Source likelihood", "Transparent description/ontology cues plus assay, biological-system, exposure, and sample-matrix priors are normalized across six competing classes. Entropy, class margin, direct-evidence count, and rule provenance expose uncertainty. Values are not calibrated causal probabilities."),
        ("Extraction phase", "Explicit description rules take precedence; otherwise parsed MolLogP/TPSA, formula polarity, lipid cues, extraction method, and analyzed phase inform LLE partition or SPE retention plausibility. This is not measured recovery."),
        ("Progenesis scoring", "https://www.nonlinear.com/progenesis/qi/v3.1/faq/identifications-scoring-algorithm.aspx"),
        ("Metabolomics Standards Initiative", "https://pubmed.ncbi.nlm.nih.gov/24039616/"),
        ("Pooled-QC guidance", "https://pmc.ncbi.nlm.nih.gov/articles/PMC5960010/"),
        ("mQACC pooled-QC practice review", "https://doi.org/10.1021/acs.analchem.3c02924"),
        ("MS2 spectral entropy", "https://www.nature.com/articles/s41592-021-01331-z"),
        ("MiMeDB source categories", "https://doi.org/10.1093/nar/gkaf1272"),
        ("MTBE extraction", "https://pubmed.ncbi.nlm.nih.gov/18281723/"),
        ("Plasma extraction comparison", "https://pubmed.ncbi.nlm.nih.gov/28000704/"),
        ("Adduct complexity", "https://pubmed.ncbi.nlm.nih.gov/35058016/"),
        ("Target-decoy context", "https://pmc.ncbi.nlm.nih.gov/articles/PMC6252074/"),
    ]
    notes.extend(("Dataset warning", warning) for warning in bundle.warnings)
    return pd.DataFrame(notes, columns=["Topic", "Method / limitation"])


def run_summary(
    bundle: AnalysisBundle,
    shortlist: pd.DataFrame,
    filter_settings: dict[str, Any],
    provenance_settings: dict[str, Any] | None,
) -> pd.DataFrame:
    provenance_settings = provenance_settings or {}
    role_counts = bundle.sample_metadata["Sample role"].value_counts() if not bundle.sample_metadata.empty else pd.Series(dtype=int)
    rows = [
        ("Generated at (UTC)", bundle.generated_at, "Reproducible run timestamp"),
        ("Accepted feature reads", len(bundle.accepted_reads), "One CM feature-level accepted read"),
        ("All identification candidates", len(bundle.ci), "Every CI candidate row scanned"),
        ("Unique accepted entities", len(bundle.selected_reads), "Grouped by Accepted Compound ID"),
        ("Repeated Accepted-ID groups", bundle.duplicate_groups, "All alternative reads retained"),
        ("Reads in repeated groups", len(bundle.duplicate_reads), "Selected and non-selected alternatives"),
        ("Extra reads collapsed", bundle.extra_reads_collapsed, "Representative matrix only"),
        ("Filtered shortlist", len(shortlist), "Current manual/preset thresholds"),
        ("Normalised sample columns", len(bundle.normalized_columns), "Retained in all merged outputs"),
        ("Raw sample columns", len(bundle.raw_columns), "Retained in all merged outputs"),
        ("Biological samples mapped", int(role_counts.get("Biological", 0)), "Used for biological detection/abundance"),
        ("Pooled-QC injections mapped", int(role_counts.get("Pooled QC", 0)), "Used for repeatability and drift diagnostics"),
        ("Blank samples mapped", int(sum(role_counts.get(role, 0) for role in ("Process blank", "Extraction blank", "Solvent blank"))), "Required for blank-contribution filtering"),
        ("MS2 spectra scanned", len(bundle.fd), "All MSP records retained"),
        ("MS2 fragment peaks scanned", len(bundle.fd_peaks), "Peak-level export retained"),
        ("Ion mode", bundle.parameters.get("ion_mode"), "User-declared context"),
        ("Search tolerance (ppm)", bundle.parameters.get("mass_tolerance"), "Caps mass-error recommendation"),
        ("QC sample regex", bundle.parameters.get("qc_pattern"), "Used for pooled-QC evidence"),
        ("Score cutoff", filter_settings.get("score"), ">="),
        ("Fragmentation cutoff", filter_settings.get("fragmentation"), ">=, with declared zero-state rule"),
        ("Permit zero Fragmentation Score", filter_settings.get("allow_zero_fragmentation"), "Explicit evidence state"),
        ("Absolute mass-error cutoff (ppm)", filter_settings.get("abs_mass_error"), "<="),
        ("Isotope Similarity cutoff", filter_settings.get("isotope"), ">="),
        ("Study type", provenance_settings.get("study_type", "not scored"), "Required for contextual source priors"),
        ("Biological system", provenance_settings.get("biological_system", "not scored"), "Human, microbial, cell-line, exposure, or environmental context"),
        ("Exposure context", provenance_settings.get("exposure_context", "not scored"), "Optional food/drug/environment/viral context modifier"),
        ("Sample matrix", provenance_settings.get("sample_matrix", "not scored"), "Required for source interpretation"),
        ("Extraction method", provenance_settings.get("extraction_method", "not scored"), "Phase context"),
        ("Analyzed phase", provenance_settings.get("analyzed_phase", "not scored"), "Phase context"),
        ("Critical interpretation", "Triage—not identification FDR", "No target-decoy results, authentic standards, or external validation labels were supplied"),
    ]
    return pd.DataFrame(rows, columns=["Measure", "Value", "Interpretation"])


def _style_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    if "00 Read Me" in workbook.sheetnames:
        workbook.active = workbook.sheetnames.index("00 Read Me")
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.sheet_view.showGridLines = False
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.fill = PatternFill("solid", fgColor=NAVY)
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            worksheet.row_dimensions[1].height = 30
        sample_rows = min(worksheet.max_row, 250)
        for column_index in range(1, worksheet.max_column + 1):
            values = [worksheet.cell(row=row, column=column_index).value for row in range(1, sample_rows + 1)]
            width = max((len(str(value)) for value in values if value is not None), default=8)
            header = str(worksheet.cell(1, column_index).value or "")
            cap = 100 if "Manuscript-ready text" in header else 70 if any(term in header for term in ("Description", "Evidence", "Rationale", "Method", "limitation", "Trace")) else 30
            worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(width + 2, 10), cap)
        if worksheet.max_row > 1:
            for row in range(2, min(worksheet.max_row, 300) + 1):
                fill = PALE_AQUA if row % 2 == 0 else "FFFFFF"
                for cell in worksheet[row]:
                    cell.fill = PatternFill("solid", fgColor=fill)
                    cell.font = Font(color=INK, size=9)
                    cell.alignment = Alignment(vertical="top", wrap_text=False)
        if worksheet.title == "14 Journal Methods":
            worksheet.column_dimensions["A"].width = 24
            worksheet.column_dimensions["B"].width = 36
            worksheet.column_dimensions["C"].width = 110
            for row_index in range(2, worksheet.max_row + 1):
                report_text = str(worksheet.cell(row=row_index, column=3).value or "")
                estimated_lines = sum(max(1, math.ceil(len(line) / 140)) for line in report_text.splitlines() or [""])
                worksheet.row_dimensions[row_index].height = min(409, max(54, estimated_lines * 12 + 12))
                for cell in worksheet[row_index]:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
        if worksheet.title == "15 Task History":
            for column, width in {"A": 28, "B": 58, "C": 15, "D": 29, "E": 29, "F": 15, "G": 58}.items():
                worksheet.column_dimensions[column].width = width
            for row_index in range(2, worksheet.max_row + 1):
                worksheet.row_dimensions[row_index].height = 32
                for cell in worksheet[row_index]:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
        worksheet.auto_filter.ref = worksheet.dimensions
        for row in worksheet.iter_rows(min_row=2, max_row=min(worksheet.max_row, 2000)):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith(("https://", "http://")):
                    cell.hyperlink = cell.value
                    cell.style = "Hyperlink"
        headers = {str(cell.value): cell.column for cell in worksheet[1]}
        for header, column_index in headers.items():
            if any(term in header for term in ("Score", "Similarity", "likelihood (%)", "CV%", "D-ratio", "detection rate")) and worksheet.max_row > 2:
                letter = get_column_letter(column_index)
                lower_is_better = any(term in header for term in ("CV%", "D-ratio", "mass error", "drift"))
                worksheet.conditional_formatting.add(
                    f"{letter}2:{letter}{worksheet.max_row}",
                    ColorScaleRule(
                        start_type="min",
                        start_color="63BE7B" if lower_is_better else "F8696B",
                        mid_type="percentile",
                        mid_value=50,
                        mid_color="FFEB84",
                        end_type="max",
                        end_color="F8696B" if lower_is_better else "63BE7B",
                    ),
                )
        for decision_header in ("Status", "Filter decision", "Analytical QC decision"):
            column_index = headers.get(decision_header)
            if not column_index:
                continue
            for row_index in range(2, min(worksheet.max_row, 5000) + 1):
                cell = worksheet.cell(row=row_index, column=column_index)
                if cell.value == "Pass":
                    cell.fill = PatternFill("solid", fgColor="DDF5F2")
                    cell.font = Font(color="146B5A", bold=True)
                elif cell.value in {"Fail", "Review"}:
                    cell.fill = PatternFill("solid", fgColor="FFF0E3" if cell.value == "Review" else "FDE8E7")
                    cell.font = Font(color="9A4E00" if cell.value == "Review" else "A61B13", bold=True)
    workbook.save(path)


def workbook_readme() -> pd.DataFrame:
    rows = [
        ("Purpose", "Auditable six-file LC–MS curation from schema validation through source/extraction interpretation."),
        ("01 Project Metadata", "User-declared assay, ion mode, acquisition, matrix, extraction, and tolerance context."),
        ("02 Sample Metadata", "Editable role/group/subject/batch/order/dilution map for every CM sample column."),
        ("03 File Audit", "Hashes, records, fields, duplicates, cross-file coverage, and structural findings."),
        ("04–05 Identity audit", "All accepted feature reads plus every alternative in repeated Accepted-ID groups."),
        ("06–07 MS2", "Spectrum-level metrics and every fragment peak with relative intensity and neutral loss."),
        ("08 Representative entities", "One deterministic representative per accepted identity with selection trace."),
        ("09 Analytical QC", "Pooled-QC, blank, detection, drift, D-ratio, and technical-replicate diagnostics."),
        ("10–12 Filtering", "Threshold method, pass/fail reasons for every entity, and the final shortlist."),
        ("13 Chemistry/source/phase", "Formula properties plus evidence-weighted source and extraction-location plausibility."),
        ("14 Journal Methods", "Dataset-specific manuscript-ready methodology, curation analysis, limitations, reporting gaps, and references."),
        ("15 Task History", "Background-processing task, status, timing, duration, and outcome audit."),
        ("Critical limitation", "Annotations and thresholds are triage evidence—not identification FDR, Level-1 confirmation, causal origin, or measured extraction recovery."),
    ]
    return pd.DataFrame(rows, columns=["Section", "Contents / interpretation"])


def export_workbook(
    bundle: AnalysisBundle,
    shortlist: pd.DataFrame,
    provenance: pd.DataFrame,
    filter_settings: dict[str, Any],
    provenance_settings: dict[str, Any] | None,
    path: str | Path,
    include_all_candidates: bool = False,
    task_history: pd.DataFrame | None = None,
    progress: Callable[[int, str], None] | None = None,
) -> Path:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    _emit(progress, 3, "Preparing entity-level filter decisions…")
    filter_decisions = evaluate_filters(bundle.selected_reads, filter_settings)
    _emit(progress, 9, "Drafting dataset-specific journal methodology and analysis…")
    report_table = journal_report_table(bundle, shortlist, provenance, filter_settings, provenance_settings)
    task_history = task_history if task_history is not None else pd.DataFrame(
        columns=["Stage", "Task", "Status", "Started", "Finished", "Duration (s)", "Outcome"]
    )
    sheets: list[tuple[str, pd.DataFrame]] = [
        ("00 Read Me", workbook_readme()),
        ("00 Run Summary", run_summary(bundle, shortlist, filter_settings, provenance_settings)),
        ("01 Project Metadata", project_metadata_frame(bundle.parameters)),
        ("02 Sample Metadata", bundle.sample_metadata),
        ("02 Metadata Requirements", metadata_requirement_table()),
        ("03 File Audit", bundle.audits),
        ("04 Accepted Reads", bundle.accepted_reads),
        ("05 Repeated ID Reads", bundle.duplicate_reads),
        ("06 MS2 Spectra", bundle.fd),
        ("07 MS2 Peaks", bundle.fd_peaks),
        ("08 Selected Entities", bundle.selected_reads),
        ("09 Analytical QC", bundle.analytical_qc),
        ("10 Threshold Method", threshold_table(bundle)),
        ("11 Filter Decisions", filter_decisions),
        ("12 Filtered Shortlist", shortlist),
        ("13 Chemistry Source Phase", provenance),
        ("14 Journal Methods", report_table),
        ("15 Task History", task_history),
        ("16 Method Notes", method_notes(bundle)),
    ]
    if include_all_candidates:
        sheets.append(("17 All CI Candidates", bundle.ci))
    sheets.extend([
        ("18 Raw CM", bundle.cm),
        ("19 Raw AM", bundle.am),
        ("20 Accepted CI", bundle.accepted_ci),
        ("21 Raw ACP", bundle.acp),
        ("22 II Summary", bundle.ii),
    ])
    with pd.ExcelWriter(target, engine="openpyxl") as writer:
        for index, (name, frame) in enumerate(sheets, 1):
            _emit(progress, 12 + int(index / len(sheets) * 66), f"Writing workbook sheet {index:,}/{len(sheets):,}: {name}…")
            safe_frame(frame).to_excel(writer, sheet_name=name[:31], index=False)
    _emit(progress, 82, "Applying workbook formatting, hyperlinks, and conditional scales…")
    _style_workbook(target)
    _emit(progress, 100, "XLSX workbook complete")
    return target


def export_complete_folder(
    bundle: AnalysisBundle,
    shortlist: pd.DataFrame,
    provenance: pd.DataFrame,
    filter_settings: dict[str, Any],
    provenance_settings: dict[str, Any] | None,
    directory: str | Path,
    include_all_candidates: bool = False,
    task_history: pd.DataFrame | None = None,
    progress: Callable[[int, str], None] | None = None,
) -> Path:
    root = Path(directory)
    root.mkdir(parents=True, exist_ok=True)
    _emit(progress, 2, "Preparing complete results package…")
    filter_decisions = evaluate_filters(bundle.selected_reads, filter_settings)
    task_history = task_history if task_history is not None else pd.DataFrame(
        columns=["Stage", "Task", "Status", "Started", "Finished", "Duration (s)", "Outcome"]
    )
    tables = [
        ("00_Project_Metadata.csv", project_metadata_frame(bundle.parameters)),
        ("01_Sample_Metadata.csv", bundle.sample_metadata),
        ("02_File_Audit.csv", bundle.audits),
        ("03_All_Accepted_Feature_Reads.csv", bundle.accepted_reads),
        ("04_Repeated_Accepted_ID_Reads.csv", bundle.duplicate_reads),
        ("05_MS2_Spectra.csv", bundle.fd),
        ("06_MS2_Fragment_Peaks.csv", bundle.fd_peaks),
        ("07_Selected_Compound_Representatives.csv", bundle.selected_reads),
        ("08_Analytical_QC.csv", bundle.analytical_qc),
        ("09_Filter_Decisions_All_Entities.csv", filter_decisions),
        ("10_Filtered_Compound_Shortlist.csv", shortlist),
        ("11_Chemistry_Source_and_Phase.csv", provenance),
        ("12_Threshold_Method.csv", threshold_table(bundle)),
        ("13_Method_Notes.csv", method_notes(bundle)),
        ("14_Task_History.csv", task_history),
    ]
    for index, (name, frame) in enumerate(tables, 1):
        _emit(progress, 3 + int(index / len(tables) * 35), f"Writing stage CSV {index:,}/{len(tables):,}: {name}…")
        export_csv(frame, root / name)
    _emit(progress, 40, "Drafting manuscript-ready methodology and curation analysis…")
    report = journal_report_markdown(bundle, shortlist, provenance, filter_settings, provenance_settings)
    (root / "JOURNAL_READY_METHODS_AND_ANALYSIS.md").write_text(report, encoding="utf-8")
    (root / "JOURNAL_READY_METHODS_AND_ANALYSIS.txt").write_text(report, encoding="utf-8")
    _emit(progress, 47, "Building the formatted multi-sheet workbook…")
    export_workbook(
        bundle,
        shortlist,
        provenance,
        filter_settings,
        provenance_settings,
        root / "LCMS_Compound_Curation_Results.xlsx",
        include_all_candidates=include_all_candidates,
        task_history=task_history,
        progress=lambda percent, message: _emit(progress, 47 + int(percent * 0.43), message),
    )
    _emit(progress, 91, "Writing the machine-readable manifest and method records…")
    source_counts = provenance["Primary source class"].value_counts().to_dict() if not provenance.empty else {}
    manifest = {
        "application_method": f"LCMS Compound Curation Workbench v{__version__}",
        "generated_at": bundle.generated_at,
        "counts": {name: {"rows": len(frame), "columns": len(frame.columns)} for name, frame in tables},
        "workbook": "LCMS_Compound_Curation_Results.xlsx",
        "journal_report": "JOURNAL_READY_METHODS_AND_ANALYSIS.md",
        "task_history": "14_Task_History.csv",
        "include_all_identification_candidates": include_all_candidates,
        "filter_settings": filter_settings,
        "provenance_settings": provenance_settings,
        "input_sha256": dict(zip(bundle.audits["Role"], bundle.audits["SHA-256"])),
        "sample_role_counts": bundle.sample_metadata["Sample role"].value_counts().to_dict(),
        "curation_outcome": {
            "representative_entities": len(bundle.selected_reads),
            "identification_evidence_passes": int(filter_decisions["Identification evidence pass"].sum()),
            "analytical_qc_passes": int(filter_decisions["Analytical QC pass"].sum()),
            "final_shortlist": len(shortlist),
            "primary_source_class_counts": source_counts,
        },
        "critical_limitations": [
            "Dataset-adaptive thresholds are not identification FDR.",
            "Source likelihoods are not calibrated causal-origin probabilities.",
            "Extraction-location likelihoods are not measured recovery.",
            "Missing QC, blank, RT, CCS, standard, or target-decoy evidence is not treated as proof of quality.",
        ],
    }
    (root / "analysis_manifest.json").write_text(json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")
    notes = method_notes(bundle)
    (root / "METHOD_AND_LIMITATIONS.txt").write_text(
        "LC-MS COMPOUND CURATION METHOD AND LIMITATIONS\n\n"
        + "\n\n".join(f"{row['Topic']}\n{row['Method / limitation']}" for _, row in notes.iterrows()),
        encoding="utf-8",
    )
    method_source = Path(__file__).resolve().parents[1] / "SCIENTIFIC_METHOD.md"
    if method_source.is_file():
        (root / "SCIENTIFIC_METHOD.md").write_text(method_source.read_text(encoding="utf-8"), encoding="utf-8")
    _emit(progress, 100, "Complete results package written successfully")
    return root
